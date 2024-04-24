import json
import pandas as pd
import os
import sys
import customtkinter as ctk
import win32com.client
from openpyxl import load_workbook
from tkinter import filedialog
from PIL import Image

references = {
    "version": "1.2.2",

    "data_file": "data.json",
    "app_config_file": "app_config.json",

    "author": "Victor Zanin Rubinec"
}   

class AtalhoWindow(ctk.CTk):
    
    def __init__(self):
        super().__init__()

        self.title("Instalação")
        
        self.atalho_label = ctk.CTkLabel(self, text="Seleicone a localização do atalho:", font=ctk.CTkFont(size=15))
        self.atalho_label.grid(row=0, column=0, padx=20, pady=20, sticky="w", columnspan=2)
        
        self.atalho_input = ctk.CTkButton(self, text="Selecionar", corner_radius=5, height=40, border_spacing=10,
            width=300, text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"), border_width=2, border_color=("gray70", "gray30"),
            anchor="w", cursor="hand2", fg_color=("gray95", "gray20"), command=self.selecionar_atalho)
        self.atalho_input.grid(row=1, column=0, padx=30, sticky="w", columnspan=2)
        
        self.confirmar_button = ctk.CTkButton(self, text="Confirmar", cursor="hand2", height=35,
            font=ctk.CTkFont(size=15, weight="bold"), command=self.confirmar_event)
        self.confirmar_button.grid(row=2, column=1, padx=20, pady=20, sticky="se")
        
        self.cancelar_button = ctk.CTkButton(self, text="Cancelar", cursor="hand2", height=35,
            font=ctk.CTkFont(size=15, weight="bold"), command=self.destroy)
        self.cancelar_button.grid(row=2, column=0, padx=20, pady=20, sticky="se")
        
    def selecionar_atalho(self):
        self.atalho = filedialog.askdirectory()
        self.atalho_input.configure(text=self.atalho)
        
    def confirmar_event(self):
        if self.atalho_input._text == "Selecionar" or not self.atalho_input._text or not self.atalho_input._text.strip():
            self.atalho_input.configure(border_color="red", fg_color="lightcoral", hover_color="lightcoral", text_color=("gray10", "gray20"))
            return
        else:
            self.atalho_input.configure(border_color=("gray70", "gray30"), bg_color=("gray95", "gray20"),
                fg_color=("gray10", "gray20"), text_color=("gray10", "gray90"))
            self.criar_atalho(self.atalho)
            
    def criar_atalho(self, diretorio):
        caminho_executavel = os.path.join(os.path.dirname(sys.executable), "Auto_CheckList.py")

        # Cria o atalho
        shell = win32com.client.Dispatch("WScript.Shell")
        atalho = shell.CreateShortcut(os.path.join(diretorio, "Auto_Checklist.lnk"))
        atalho.TargetPath = caminho_executavel
        atalho.WorkingDirectory = os.path.dirname(caminho_executavel)
        atalho.save()

        print(f"Atalho para 'programa' criado em '{os.path.join(diretorio, 'Auto_Checklist.lnk')}'")

        # Fecha a janela após criar o atalho
        self.destroy()
       
class App(ctk.CTk):
        
    def __init__(self):
        super().__init__()

        self.title("Auto CheckList")
        self.geometry("650x400")
        
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        
        def get_assets_path():
            if getattr(sys, 'frozen', False):
                # Se o programa estiver empacotado como um executável
                return os.path.join(os.path.dirname(sys.executable), "assets")
            else:
                # Se estiver em modo de desenvolvimento
                return os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")
        
        image_path = get_assets_path()
        self.logo_image = ctk.CTkImage(Image.open(os.path.join(image_path, "logo_light.png")), 
            dark_image=Image.open(os.path.join(image_path, "logo_dark.png")), size=(130, 50))
        self.home_image = ctk.CTkImage(Image.open(os.path.join(image_path, "home_light.png")),
            dark_image=Image.open(os.path.join(image_path, "home_dark.png")), size=(20, 20))
        self.ajuda_image = ctk.CTkImage(Image.open(os.path.join(image_path, "ajuda_light.png")),
            dark_image=Image.open(os.path.join(image_path, "ajuda_dark.png")), size=(20, 20))
        self.configuracao_image = ctk.CTkImage(Image.open(os.path.join(image_path, "configuracao_light.png")),
            dark_image=Image.open(os.path.join(image_path, "configuracao_dark.png")), size=(20, 20))

        self.navegacao_frame = ctk.CTkFrame(self, corner_radius=0, width=150)
        self.navegacao_frame.grid(row=0, column=0, sticky="nsew")
        self.navegacao_frame.grid_rowconfigure(4, weight=1)

        self.navegacao_frame_label = ctk.CTkLabel(self.navegacao_frame, text="Auto CheckList",
            compound="left", font=ctk.CTkFont(size=20, weight="bold"))
        self.navegacao_frame_label.grid(row=0, column=0, padx=20, pady=20)
        
        self.home_button = ctk.CTkButton(self.navegacao_frame, corner_radius=0, height=40, border_spacing=10, text="Home",
            fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
            image=self.home_image, anchor="w", command=self.home_button_event)
        self.home_button.grid(row=1, column=0, sticky="ew")

        self.ajuda_button = ctk.CTkButton(self.navegacao_frame, corner_radius=0, height=40, border_spacing=10, text="Ajuda",
            fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
            image=self.ajuda_image, anchor="w", command=self.ajuda_button_event)
        self.ajuda_button.grid(row=2, column=0, sticky="ew")

        self.configuracao_button = ctk.CTkButton(self.navegacao_frame, corner_radius=0, height=40, border_spacing=10, text="Configuração",
            fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
            image=self.configuracao_image, anchor="w", command=self.configuracao_button_event)
        self.configuracao_button.grid(row=3, column=0, sticky="ew")
        
        self.logo_position = ctk.CTkLabel(self.navegacao_frame, image=self.logo_image, compound="left", text="")
        self.logo_position.grid(row=6, column=0, sticky="s",  pady=10)
        
        # Frame home
        self.home_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent", width=450)
        self.home_frame.grid_columnconfigure(1, weight=1)
        self.home_frame.grid_rowconfigure(5, weight=1)
        
        self.cnpj_label = ctk.CTkLabel(self.home_frame, text="Digite o CPNJ desejado:", font=ctk.CTkFont(size=15))
        self.cnpj_label.grid(row=0, column=0, padx=20, pady=15, sticky="w")

        self.cnpj_input = ctk.CTkEntry(self.home_frame, font=ctk.CTkFont(size=15), width=400, height=40, 
            placeholder_text="Digite o CNPJ aqui...")
        self.cnpj_input.grid(row=1, column=0, padx=30, sticky="w")
        
        self.arquivo_label = ctk.CTkLabel(self.home_frame, text="Selecione o arquivo:", font=ctk.CTkFont(size=15))
        self.arquivo_label.grid(row=2, column=0, padx=20, pady=15, sticky="w")
        
        self.arquivo_input = ctk.CTkButton(self.home_frame, text=app_config["arquivo"], corner_radius=5, height=40, border_spacing=10,
            width=400, text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"), border_width=2, border_color=("gray70", "gray30"),
            anchor="w", cursor="hand2", fg_color=("gray95", "gray20"), command=self.selecionar_arquivo)
        self.arquivo_input.grid(row=3, column=0, padx=30, sticky="w")
        
        if self.arquivo_input._text == "":
            self.arquivo_input.configure(text="Clique aqui para selecionar um arquivo...")
            
        self.error_label = ctk.CTkLabel(self.home_frame, text="", font=ctk.CTkFont(size=12, weight="bold"), text_color="lightcoral",
            wraplength=400, justify="left")
        self.error_label.grid(row=4, column=0, padx=20, pady=5, sticky="w")
        
        self.confirmar_button = ctk.CTkButton(self.home_frame, text="Confirmar", cursor="hand2", height=35, 
            font=ctk.CTkFont(size=15, weight="bold"), command=self.verificar_componentes)
        self.confirmar_button.grid(row=6, column=0, padx=20, pady=20, sticky="se")
        
        # Frame ajuda
        self.ajuda_frame = ctk.CTkScrollableFrame(self, corner_radius=0, fg_color="transparent", width=450)
        self.ajuda_frame.grid_columnconfigure(0, weight=1)
        self.ajuda_frame.grid_rowconfigure(10, weight=1)
        
        self.ajuda_titulo = ctk.CTkLabel(self.ajuda_frame, text="Bem-vindo à seção de \nAjuda do Auto CheckList!",
            font=ctk.CTkFont(size=20, weight="bold"))
        self.ajuda_titulo.grid(row=0, column=0, padx=15, pady=10, sticky="n")
        
        self.ajuda_texto = ctk.CTkLabel(self.ajuda_frame, text="  Este programa foi projetado para facilitar o preenchimento automático de planilhas, economizando tempo e minimizando erros. Aqui estão algumas dicas úteis para utilizar o programa:",
            font=ctk.CTkFont(size=12), wraplength=430, justify="left")
        self.ajuda_texto.grid(row=1, column=0, padx=15, pady=5, sticky="w")
        
        self.ajuda_titulo_topico1 = ctk.CTkLabel(self.ajuda_frame, text="1. Selecionar Arquivo Checklist:", 
            font=ctk.CTkFont(size=15, weight="bold"))  
        self.ajuda_titulo_topico1.grid(row=2, column=0, padx=10, pady=5, sticky="w")
        
        self.ajuda_texto_topico1 = ctk.CTkLabel(self.ajuda_frame, text="  Certifique-se de selecionar o arquivo checklist correto na tela inicial. O programa só aceita arquivos no formato de checklist.",
            font=ctk.CTkFont(size=12), wraplength=430, justify="left")
        self.ajuda_texto_topico1.grid(row=3, column=0, padx=15, pady=5, sticky="w")
        
        self.ajuda_titulo_topico2 = ctk.CTkLabel(self.ajuda_frame, text="2. Inserir CNPJ:", 
            font=ctk.CTkFont(size=15, weight="bold"))  
        self.ajuda_titulo_topico2.grid(row=4, column=0, padx=10, pady=5, sticky="w")
        
        self.ajuda_texto_topico2 = ctk.CTkLabel(self.ajuda_frame, text="  Digite o CNPJ desejado no campo apropriado. Lembre-se de que o CNPJ deve ter 14 dígitos numéricos.",
            font=ctk.CTkFont(size=12), wraplength=430, justify="left")
        self.ajuda_texto_topico2.grid(row=5, column=0, padx=15, pady=5, sticky="w")

        self.ajuda_titulo_topico3 = ctk.CTkLabel(self.ajuda_frame, text="3. Verificar Campos da Planilha:", 
            font=ctk.CTkFont(size=15, weight="bold"))  
        self.ajuda_titulo_topico3.grid(row=6, column=0, padx=10, pady=5, sticky="w")
        
        self.ajuda_texto_topico3 = ctk.CTkLabel(self.ajuda_frame, text="  Antes de confirmar, verifique se os campos da planilha estão vazios. O programa não irá sobrescrever dados existentes.",
            font=ctk.CTkFont(size=12), wraplength=430, justify="left")
        self.ajuda_texto_topico3.grid(row=7, column=0, padx=15, pady=5, sticky="w")

        self.ajuda_titulo_topico4 = ctk.CTkLabel(self.ajuda_frame, text="4. Confirmação:", 
            font=ctk.CTkFont(size=15, weight="bold"))  
        self.ajuda_titulo_topico4.grid(row=8, column=0, padx=10, pady=5, sticky="w")
        
        self.ajuda_texto_topico4 = ctk.CTkLabel(self.ajuda_frame, text="  Após seguir essas etapas, clique no botão 'Confirmar' para inserir os dados na planilha.",
            font=ctk.CTkFont(size=12), wraplength=430, justify="left")
        self.ajuda_texto_topico4.grid(row=9, column=0, padx=15, pady=5, sticky="w")

        self.ajuda_texto_final = ctk.CTkLabel(self.ajuda_frame, text="  Caso tenha alguma dúvida adicional, sinta-se à vontade para entrar em contato conosco.\n\n  Obrigado por utilizar o Auto CheckList!",
            font=ctk.CTkFont(size=12), wraplength=430, justify="left")
        self.ajuda_texto_final.grid(row=10, column=0, padx=15, pady=5, sticky="w")
        
        # Frame configuração
        self.configuracao_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent", width=450)
        self.configuracao_frame.grid_columnconfigure(0, weight=1)
        self.configuracao_frame.grid_rowconfigure(4, weight=1)
        
        self.appearance_mode_label = ctk.CTkLabel(self.configuracao_frame, text="Tema:", font=ctk.CTkFont(size=15))
        self.appearance_mode_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")
        
        self.appearance_mode_menu = ctk.CTkOptionMenu(self.configuracao_frame, values=["System", "Light", "Dark"],
            command=self.change_appearance_mode_event)
        self.appearance_mode_menu.grid(row=1, column=0, padx=20, sticky="w")
        
        self.versao_label = ctk.CTkLabel(self.configuracao_frame, text=f"Versão: {references['version']}",
            font=ctk.CTkFont(size=12), text_color=("gray50", "gray30"))
        self.versao_label.grid(row=6, column=0, padx=10, pady=10, sticky="w")
        
        self.autor_label = ctk.CTkLabel(self.configuracao_frame, text=f"Desenvolvido por {references['author']}", 
            font=ctk.CTkFont(size=12), text_color=("gray50", "gray30"))
        self.autor_label.grid(row=6, column=1, padx=10, pady=10, sticky="w")
        
        self.selecionar_frame_por_nome("home")
    
    def buscar_por_cnpj(self, df, cnpj):
        cnpj = cnpj.replace('.', '').replace('/', '').replace('-', '')
        for cliente in df.iterrows():
            if cnpj in cliente[1]["CNPJ"]:
                valores = cliente[1]["CNPJ"].split(";")
                for valor in valores:
                    if cnpj in valor:
                        cliente_encontrado = cliente[1].copy() 
                        cliente_encontrado['CNPJ'] = valor 
                        return pd.DataFrame([cliente_encontrado])
        return pd.DataFrame()
        
    def verificar_componentes(self):
        self.error_label.configure(text="", text_color="lightcoral")
        if not all([app_config['diretorio'], app_config['arquivo']]):
            self.error_label.configure(text="Selecione um arquivo!")
            self.arquivo_input.configure(border_color="red", fg_color="lightcoral", hover_color="lightcoral", text_color=("gray10", "gray20"))
            return
        elif not self.cnpj_input.get():
            self.error_label.configure(text="Digite um CNPJ!")
            self.cnpj_input.configure(border_color="red", fg_color="lightcoral",
                text_color=("gray10", "gray20"))
            self.cnpj_input.focus()
            return
        elif len(self.cnpj_input.get()) != 14 or not self.cnpj_input.get().isdigit():
            self.error_label.configure(text="CNPJ inválido!")
            self.cnpj_input.configure(border_color="red", fg_color="lightcoral",
                text_color=("gray10", "gray20"))
            self.cnpj_input.focus()
            return
        elif not os.path.exists(f"{app_config['diretorio']}/{app_config['arquivo']}"):
            self.error_label.configure(text="Arquivo não encontrado!")
            self.arquivo_input.configure(border_color="red", fg_color="lightcoral", hover_color="lightcoral", text_color=("gray10", "gray20"))
            return
        elif not self.verificar_planilha(f"{app_config['diretorio']}/{app_config['arquivo']}"):
            self.error_label.configure(text="Planilha não possui a aba 'Checklist - Vendas'")
            self.arquivo_input.configure(border_color="red", fg_color="lightcoral", hover_color="lightcoral", text_color=("gray10", "gray20"))
            return
        elif not self.verificar_campo():
            self.error_label.configure(text="Campos da planilha não estão vazios!")
            self.arquivo_input.configure(border_color="red", fg_color="lightcoral", hover_color="lightcoral", text_color=("gray10", "gray20"))
            return
        else:
            self.cnpj_input.configure(border_color=("gray70", "gray30"), bg_color=("gray95", "gray20"),
                fg_color=("gray10", "gray20"), text_color=("gray10", "gray90"))
            self.arquivo_input.configure(border_color=("gray70", "gray30"), bg_color=("gray95", "gray20"),
                fg_color=("gray10", "gray20"), hover_color=("gray70", "gray30"), text_color=("gray10", "gray90"))
            self.error_label.configure(text="")

            cliente = self.buscar_por_cnpj(df, self.cnpj_input.get())
            self.confirmar(cliente)
            
    def verificar_planilha(self, diretorio):
        try:
            workbook = load_workbook(diretorio)
            return "Checklist - Vendas" in workbook.sheetnames
        except Exception as e:
            return False

    def verificar_campo(self):
        try:
            workbook = load_workbook(f"{app_config['diretorio']}/{app_config['arquivo']}")
            sheet = workbook["Checklist - Vendas"]
            celulas = ['C4', 'G4', 'G5', 'G6', 'G7', 'H8']
            for cell in celulas:
                if sheet[cell].value != None:
                    return False
            return True
        except Exception as e:
            return False
    
    def confirmar(self, cliente):
        arquivo = load_workbook(f"{app_config['diretorio']}/{app_config['arquivo']}") 
        planilha = arquivo['Checklist - Vendas']
        
        if cliente.empty:
            self.error_label.configure(text="CNPJ não encontrado!")
            self.cnpj_input.configure(border_color="red", fg_color="lightcoral",
                text_color=("gray10", "gray20"))
            self.cnpj_input.focus()
            return
        elif len(cliente) > 1:
            self.error_label.configure(text="CNPJ duplicado!")
            self.cnpj_input.configure(border_color="red", fg_color="lightcoral",
                text_color=("gray10", "gray20"))
            self.cnpj_input.focus()
            return
        else:
            self.error_label.configure(text="")
            self.cnpj_input.configure(border_color=("gray70", "gray30"), bg_color=("gray95", "gray20"),
                fg_color=("gray10", "gray20"), text_color=("gray10", "gray90"))
            
            # atualizar os valores
            planilha['C4'] = cliente['CNPJ'].values[0]
            planilha['G4'] = cliente['nome_cliente'].values[0].title()
            planilha['G5'] = cliente['endereco_fisico'].values[0].title()
            planilha['G6'] = cliente['estado'].values[0].title()
            planilha['G7'] = cliente['cidade'].values[0].title()
            planilha['H8'] = cliente['party_id'].values[0]
            
            self.error_label.configure(text="Dados inseridos com sucesso!", text_color="green")
            
            # atualizar o arquivo
            arquivo.save(f"{app_config['diretorio']}/{app_config['arquivo']}")
            
    def selecionar_arquivo(self):
        self.error_label.configure(text="", text_color="lightcoral")
        diretorio_saida = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
        if diretorio_saida:
            resultado = diretorio_saida.split("/")
            diretorio = "/".join(resultado[:-1])
            arquivo = resultado[-1]
            if self.verificar_planilha(diretorio_saida):
                self.arquivo_input.configure(text=arquivo)
                app_config['diretorio'] = diretorio
                app_config['arquivo'] = arquivo
                json.dump(app_config, open(app_config_file, 'w'), indent=4)
                
                self.arquivo_input.configure(border_color=("gray70", "gray30"), bg_color=("gray95", "gray20"),
                    fg_color=("gray10", "gray20"), hover_color=("gray70", "gray30"), text_color=("gray10", "gray90"))
                self.error_label.configure(text="")
            else:
                self.error_label.configure(text="Planilha não possui a aba 'Checklist - Vendas', selecione outro arquivo!")
                return
                
    def selecionar_frame_por_nome(self, name):
        self.home_button.configure(fg_color=("gray75", "gray25") if name == "home" else "transparent")
        self.ajuda_button.configure(fg_color=("gray75", "gray25") if name == "ajuda" else "transparent")
        self.configuracao_button.configure(fg_color=("gray75", "gray25") if name == "configuracao" else "transparent")

        if name == "home":
            self.home_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.home_frame.grid_forget()
        if name == "ajuda":
            self.ajuda_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.ajuda_frame.grid_forget()
        if name == "configuracao":
            self.configuracao_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.configuracao_frame.grid_forget()        

    def home_button_event(self):
        self.selecionar_frame_por_nome("home")

    def ajuda_button_event(self):
        self.selecionar_frame_por_nome("ajuda")

    def configuracao_button_event(self):
        self.selecionar_frame_por_nome("configuracao")
        
    def change_appearance_mode_event(self, new_appearance_mode):
        ctk.set_appearance_mode(new_appearance_mode)
        
if __name__ == "__main__":
    
    def get_config_path():
        if getattr(sys, 'frozen', False):
            # Se o programa estiver empacotado como um executável
            return os.path.join(os.path.dirname(sys.executable), "config")
        else:
            # Se estiver em modo de desenvolvimento
            return os.path.join(os.path.dirname(os.path.abspath(__file__)), "config")            
                
    app_config_model = {
        "diretorio": "",
        "arquivo": ""
    }
    
    config_path = get_config_path()
    
    app_config_file = os.path.join(config_path, "app_config.json")
    data_file = os.path.join(config_path, "data.json")
    
    if getattr(sys, 'frozen', False):
        if not os.path.exists(app_config_file):
            
            atalho_window = AtalhoWindow()
            atalho_window.mainloop()
            
            with open(app_config_file, 'w') as f:
                json.dump(app_config_model, f, indent=4)
            app_config = json.load(open(app_config_file))
        else:
            app_config = json.load(open(app_config_file))
    else:
        with open(app_config_file, 'w') as f:
            json.dump(app_config_model, f, indent=4)
        app_config = json.load(open(app_config_file))

    data = json.load(open(data_file))
    df = pd.DataFrame(data['cliente'])

    app = App()
    app.mainloop()
