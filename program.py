import json
import pandas as pd
import os
from openpyxl import load_workbook
from tkinter import messagebox
from tkinter import filedialog
import customtkinter as ctk
from PIL import Image
    
class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Auto Faturamento")
        self.geometry("700x400")
        
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        
        # Puxando imagens do assets
        image_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "assets")
        self.logo_image = ctk.CTkImage(Image.open(os.path.join(image_path, "logo_light.png")), 
            dark_image=Image.open(os.path.join(image_path, "logo_dark.png")), size=(130, 50))
        self.home_image = ctk.CTkImage(Image.open(os.path.join(image_path, "home_light.png")),
            dark_image=Image.open(os.path.join(image_path, "home_dark.png")), size=(20, 20))
        self.ajuda_image = ctk.CTkImage(Image.open(os.path.join(image_path, "ajuda_light.png")),
            dark_image=Image.open(os.path.join(image_path, "ajuda_dark.png")), size=(20, 20))
        self.configuracao_image = ctk.CTkImage(Image.open(os.path.join(image_path, "configuracao_light.png")),
            dark_image=Image.open(os.path.join(image_path, "configuracao_dark.png")), size=(20, 20))
        
        
        self.navigation_frame = ctk.CTkFrame(self, corner_radius=0)
        self.navigation_frame.grid(row=0, column=0, sticky="nsew")
        self.navigation_frame.grid_rowconfigure(4, weight=1)

        self.navigation_frame_label = ctk.CTkLabel(self.navigation_frame, text="Auto Faturamento",
            compound="left", font=ctk.CTkFont(size=20, weight="bold"))
        self.navigation_frame_label.grid(row=0, column=0, padx=20, pady=20)
        
        self.home_button = ctk.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Home",
            fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
            image=self.home_image, anchor="w", command=self.home_button_event)
        self.home_button.grid(row=1, column=0, sticky="ew")

        self.ajuda_button = ctk.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Ajuda",
            fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
            image=self.ajuda_image, anchor="w", command=self.ajuda_button_event)
        self.ajuda_button.grid(row=2, column=0, sticky="ew")

        self.configuracao_button = ctk.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Configuração",
            fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
            image=self.configuracao_image, anchor="w", command=self.configuracao_button_event)
        self.configuracao_button.grid(row=3, column=0, sticky="ew")
        
        self.logo_position = ctk.CTkLabel(self.navigation_frame, image=self.logo_image, compound="left", text="")
        self.logo_position.grid(row=6, column=0, sticky="s",  pady=10)
        
        # Frame home
        self.home_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.home_frame.grid_columnconfigure(1, weight=1)
        self.home_frame.grid_rowconfigure(4, weight=1)
        
        # Componentes do frame home
        self.cnpj_label = ctk.CTkLabel(self.home_frame, text="Digite o CPNJ desejado:", font=ctk.CTkFont(size=15))
        self.cnpj_label.grid(row=0, column=0, padx=20, pady=15, sticky="w")

        self.cnpj_input = ctk.CTkEntry(self.home_frame, font=ctk.CTkFont(size=15), width=400, height=40)
        self.cnpj_input.grid(row=1, column=0, padx=30, sticky="e")
        
        self.arquivo_label = ctk.CTkLabel(self.home_frame, text="Selecione o arquivo:", font=ctk.CTkFont(size=15))
        self.arquivo_label.grid(row=2, column=0, padx=20, pady=15, sticky="w")
        
        self.arquivo_input = ctk.CTkButton(self.home_frame, text=app_config["arquivo"], corner_radius=5, height=40, border_spacing=10,
            width=400, text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"), border_width=2, border_color=("gray70", "gray30"),
            anchor="w", cursor="hand2", fg_color=("gray95", "gray20"), command=self.selecionar_arquivo)
        self.arquivo_input.grid(row=3, column=0, padx=30, sticky="w")
        
        self.confirmar_button = ctk.CTkButton(self.home_frame, text="Confirmar", cursor="hand2", height=35, command=self.verificar_componentes)
        self.confirmar_button.grid(row=6, column=0, padx=20, pady=20, sticky="se")
        
        # Frame ajuda
        self.ajuda_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")

        # Frame configuração
        self.configuracao_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        
        self.appearance_mode_menu = ctk.CTkOptionMenu(self.configuracao_frame, values=["System", "Dark", "Light"],
            command=self.change_appearance_mode_event)
        self.appearance_mode_menu.grid(row=0, column=0, padx=20, pady=20, sticky="s")
        
        self.select_frame_by_name("home")
        
    def buscar_por_cnpj(self, df, cnpj):
        cnpj = cnpj.replace('.', '').replace('/', '').replace('-', '')
        for cliente in df.iterrows():
            if cnpj in cliente[1]["CNPJ"]:
                valores = cliente[1]["CNPJ"].split(";")
                for valor in valores:
                    if cnpj in valor:
                        cliente_encontrado = cliente[1].copy() 
                        cliente_encontrado['CNPJ'] = valor 
                        return pd.DataFrame([cliente_encontrado])
        return pd.DataFrame()
        
    def verificar_componentes(self):
        if not all([app_config['diretorio'], app_config['arquivo']]):
            messagebox.showwarning("Aviso", "Selecione um arquivo para continuar.")
            return
        elif not self.cnpj_input.get():
            messagebox.showwarning("Aviso", "Digite um CNPJ para continuar.")
            return
        elif len(self.cnpj_input.get()) != 14 or not self.cnpj_input.get().isdigit():
            messagebox.showwarning("Aviso", "CNPJ inválido.")
            return
        elif not os.path.exists(os.path.join(app_config['diretorio'], app_config['arquivo'])):
            messagebox.showwarning("Aviso", "Arquivo não encontrado.")
            return
        elif not self.verificar_planilha(os.path.join(app_config['diretorio'], app_config['arquivo'])):
            messagebox.showwarning("Aviso", f"Planilha não possui a aba 'Checklist - Vendas'")
            return
        elif not self.verificar_campo():
            messagebox.showwarning("Aviso", "Planilha possui campos preenchidos.")
            return

        cliente = self.buscar_por_cnpj(df, self.cnpj_input.get())
        self.confirmar(cliente)
            
    def verificar_planilha(self, diretorio):
        try:
            workbook = load_workbook(diretorio)
            return "Checklist - Vendas" in workbook.sheetnames
        except Exception as e:
            print(f"Ocorreu um erro ao tentar verificar a planilha: {e}")
            return False

    def verificar_campo(self):
        try:
            workbook = load_workbook(os.path.join(app_config['diretorio'], app_config['arquivo']))
            sheet = workbook["Checklist - Vendas"]
            celulas = ['C4', 'G4', 'G5', 'G6', 'G7', 'H8']
            for cell in celulas:
                print(sheet[cell].value)
                if sheet[cell].value != None:
                    return False
            return True
        except Exception as e:
            print(f"Ocorreu um erro ao tentar verificar os campos: {e}")
            return False
    
    def confirmar(self, cliente):
        arquivo = load_workbook(f"{app_config['diretorio']}/{app_config['arquivo']}") 
        planilha = arquivo['Checklist - Vendas']
        
        if cliente.empty:
            messagebox.showwarning("Aviso", "Cliente não encontrado.")
            return
        elif len(cliente) > 1:
            messagebox.showwarning("Aviso", "Cliente duplicado.")
            return
        else:
            # atualizar os valores
            planilha['C4'] = cliente['CNPJ'].values[0]
            planilha['G4'] = cliente['nome_cliente'].values[0].title()
            planilha['G5'] = cliente['endereco_fisico'].values[0].title()
            planilha['G6'] = cliente['estado'].values[0].title()
            planilha['G7'] = cliente['cidade'].values[0].title()
            planilha['H8'] = cliente['party_id'].values[0]
            
            # atualizar o arquivo
            arquivo.save(f"{app_config['diretorio']}/{app_config['arquivo']}")
            print("Cliente atualizado com sucesso.")
            
    def selecionar_arquivo(self):
        diretorio_saida = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
        if diretorio_saida:
            resultado = diretorio_saida.split("/")
            diretorio = "/".join(resultado[:-1])
            arquivo = resultado[-1]
            if self.verificar_planilha(diretorio_saida):
                self.arquivo_input.configure(text=arquivo)
                app_config['diretorio'] = diretorio
                app_config['arquivo'] = arquivo
                json.dump(app_config, open("app_config.json", 'w'), indent=4)
        
    def select_frame_by_name(self, name):
        self.home_button.configure(fg_color=("gray75", "gray25") if name == "home" else "transparent")
        self.ajuda_button.configure(fg_color=("gray75", "gray25") if name == "ajuda" else "transparent")
        self.configuracao_button.configure(fg_color=("gray75", "gray25") if name == "configuracao" else "transparent")

        if name == "home":
            self.home_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.home_frame.grid_forget()
        if name == "ajuda":
            self.ajuda_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.ajuda_frame.grid_forget()
        if name == "configuracao":
            self.configuracao_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.configuracao_frame.grid_forget()

    def home_button_event(self):
        self.select_frame_by_name("home")

    def ajuda_button_event(self):
        self.select_frame_by_name("ajuda")

    def configuracao_button_event(self):
        self.select_frame_by_name("configuracao")
        
    def change_appearance_mode_event(self, new_appearance_mode):
        ctk.set_appearance_mode(new_appearance_mode)
        
if __name__ == "__main__":
    app_config = {
        "diretorio": "",
        "arquivo": ""
    }
    if not os.path.exists("app_config.json"):
        with open("app_config.json", 'w') as f:
            json.dump(app_config, f, indent=4)
    else:
        app_config = json.load(open("app_config.json"))

    data = json.load(open("data.json"))
    df = pd.DataFrame(data['cliente'])

    app = App()
    app.mainloop()
